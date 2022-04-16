# no need to install anything
import os
import sys
# pip install pyqt5, pip install pyqt5 tools
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtGui import QPixmap
# just change the name
from ID import Ui_MainWindow

from openpyxl import load_workbook
import numpy as np

wb = load_workbook('database.xlsx')


class MainWindow:
    def __init__(self):
        # the way app working
        self.main_win = QMainWindow()
        self.uic = Ui_MainWindow()
        self.uic.setupUi(self.main_win)
        # khai bao nut an
        self.uic.Button_find.clicked.connect(self.showinfo)

    def showinfo(self):
        a = self.uic.Screen_ID.toPlainText()
        print(a)
        # in ra gia tri trong mot cot////
        ws = wb['Tom']
        col = ws['b']
        col_value = []
        for i in col:
            col_value.append(i.value)
        try:
            print(col_value)
            Row_number = col_value.index(a)
            print(Row_number)

            # in ra gia tri trong mot hang////
            ws = wb['Tom']
            row = ws[Row_number + 1]
            row_value = []
            for i in row:
                row_value.append(i.value)
            print(row_value)
            self.uic.Screen_name.setText(row_value[2])
            self.uic.Screen_age.setText(str(row_value[3]))

            try:
                # show pic
                link = 'Pictures/'
                filename = os.listdir(link)
                print(filename)
                self.uic.Screen_pic.setPixmap(QPixmap(f'{link}{filename[Row_number - 1]}'))
                # self.uic.Screen_pic.setPixmap(QPixmap('Pictures/1A.jpg'))
            except:
                self.uic.Screen_pic.setText("khong co hinh")
        except:
            self.uic.Screen_name.setText("khong co ai")

    def show(self):
        # command to run
        self.main_win.show()


if __name__ == "__main__":
    # run app
    app = QApplication(sys.argv)
    main_win = MainWindow()
    main_win.show()
    sys.exit(app.exec())
